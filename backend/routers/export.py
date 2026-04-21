from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from database import get_db
import models, sizing
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

router = APIRouter(prefix="/customers/{customer_id}/export", tags=["export"])

BLUE = "0070C0"
DARK_BLUE = "003366"
YELLOW = "FFFF00"
GREEN = "70AD47"
LIGHT_GRAY = "F2F2F2"
ORANGE = "FFA500"
WHITE = "FFFFFF"

thin = Side(style="thin", color="AAAAAA")
thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)


def _hdr(ws, row, col, text, bg=DARK_BLUE, fg=WHITE, bold=True, wrap=False, size=11):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(bold=bold, color=fg, size=size)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    cell.border = thin_border
    return cell


def _cell(ws, row, col, value, bold=False, bg=None, align="left", number_format=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, size=10)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    if number_format:
        cell.number_format = number_format
    cell.border = thin_border
    return cell


def _section(ws, row, col, text, span=8):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(bold=True, color=WHITE, size=11)
    cell.fill = PatternFill("solid", fgColor=BLUE)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + span - 1)
    return cell


def build_excel(customer_id: int, db: Session):
    c = db.query(models.Customer).filter(models.Customer.id == customer_id).first()
    if not c:
        raise HTTPException(status_code=404, detail="Customer not found")

    ws_data = db.query(models.WorkloadSurvey).filter(models.WorkloadSurvey.customer_id == customer_id).first()
    items = db.query(models.WorkloadItem).filter(models.WorkloadItem.survey_id == ws_data.id).all() if ws_data else []
    net = db.query(models.NetworkSurvey).filter(models.NetworkSurvey.customer_id == customer_id).first()
    bk = db.query(models.BackupSurvey).filter(models.BackupSurvey.customer_id == customer_id).first()
    sources = db.query(models.BackupSource).filter(models.BackupSource.survey_id == bk.id).all() if bk else []
    inv = db.query(models.PhysicalInventory).filter(models.PhysicalInventory.customer_id == customer_id).first()
    app_inv = db.query(models.ApplicationInventory).filter(models.ApplicationInventory.customer_id == customer_id).first()
    sec = db.query(models.SecuritySurvey).filter(models.SecuritySurvey.customer_id == customer_id).first()
    ocp = db.query(models.OCPSurvey).filter(models.OCPSurvey.customer_id == customer_id).first()

    wb = Workbook()

    # ── Sheet 1: Cover ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "THÔNG TIN CHUNG"
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 45

    ws.merge_cells("A1:B1")
    title = ws["A1"]
    title.value = "INFRASTRUCTURE SIZING – CUSTOMER SURVEY REPORT"
    title.font = Font(bold=True, size=16, color=WHITE)
    title.fill = PatternFill("solid", fgColor=DARK_BLUE)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:B2")
    sub = ws["A2"]
    sub.value = f"Ngày xuất báo cáo: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    sub.font = Font(italic=True, size=10)
    sub.alignment = Alignment(horizontal="center")

    fields = [
        ("Khách hàng / Customer", c.name),
        ("Dự án / Project Name", c.project_name or ""),
        ("Người phụ trách / Contact", c.contact or ""),
        ("Email", c.email or ""),
        ("Điện thoại", c.phone or ""),
        ("Ngày khảo sát / Survey Date", c.survey_date or ""),
        ("Presales phụ trách", c.presales or ""),
        ("Ghi chú", c.notes or ""),
    ]
    for i, (label, value) in enumerate(fields, start=3):
        _cell(ws, i, 1, label, bold=True, bg=LIGHT_GRAY)
        _cell(ws, i, 2, value)

    # ── Sheet 2: Workload Survey ────────────────────────────────────────────
    ws2 = wb.create_sheet("WORKLOAD SURVEY")
    ws2.column_dimensions["A"].width = 5
    ws2.column_dimensions["B"].width = 30
    ws2.column_dimensions["C"].width = 20
    for col in "DEFGHIJKLMN":
        ws2.column_dimensions[col].width = 14

    ws2.merge_cells("A1:N1")
    ws2["A1"].value = "WORKLOAD SURVEY – THÔNG TIN TÀI NGUYÊN WORKLOAD"
    ws2["A1"].font = Font(bold=True, size=14, color=WHITE)
    ws2["A1"].fill = PatternFill("solid", fgColor=DARK_BLUE)
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 35

    if ws_data:
        _section(ws2, 2, 1, "PHẦN A – THÔNG TIN MÔI TRƯỜNG TỔNG QUAN", 14)
        env_rows = [
            ("Loại môi trường", ws_data.env_type),
            ("Nền tảng ảo hóa", ws_data.virt_platform),
            ("Hệ thống hiện tại", ws_data.current_system),
            ("Yêu cầu HA", "Yes" if ws_data.ha_required else "No"),
            ("Cluster nodes dự kiến", ws_data.cluster_nodes),
            ("Thời gian dự phòng tăng trưởng (năm)", ws_data.growth_years),
            ("Tỉ lệ tăng trưởng dự kiến / năm (%)", ws_data.growth_rate),
        ]
        for i, (label, val) in enumerate(env_rows, start=3):
            _cell(ws2, i, 1, "", bg=LIGHT_GRAY)
            ws2.merge_cells(start_row=i, start_column=2, end_row=i, end_column=7)
            _cell(ws2, i, 2, label, bold=True, bg=LIGHT_GRAY)
            ws2.merge_cells(start_row=i, start_column=8, end_row=i, end_column=14)
            _cell(ws2, i, 8, val)

        row = 10
        _section(ws2, row, 1, "PHẦN B – CHI TIẾT WORKLOAD / VM GROUPS", 14)
        row += 1
        headers = ["STT", "Tên Workload", "Loại Workload", "Số VM", "vCPU/VM", "RAM (GB)/VM",
                   "Disk OS (GB)/VM", "Disk Data (GB)/VM", "IOPS/VM", "Throughput (MB/s)/VM",
                   "OS Type", "Tier", "Ghi chú"]
        for j, h in enumerate(headers, start=1):
            _hdr(ws2, row, j, h, bg=BLUE, size=9, wrap=True)
        ws2.row_dimensions[row].height = 35

        total_vcpu = total_ram = total_os = total_data = 0
        for item in items:
            row += 1
            n = item.vm_count or 0
            total_vcpu += n * (item.vcpu_per_vm or 0)
            total_ram += n * (item.ram_gb_per_vm or 0)
            total_os += n * (item.disk_os_gb_per_vm or 0)
            total_data += n * (item.disk_data_gb_per_vm or 0)
            vals = [item.order_no, item.name, item.workload_type, n, item.vcpu_per_vm,
                    item.ram_gb_per_vm, item.disk_os_gb_per_vm, item.disk_data_gb_per_vm,
                    item.iops_per_vm, item.throughput_mbps_per_vm, item.os_type, item.tier, item.notes]
            for j, v in enumerate(vals, start=1):
                _cell(ws2, row, j, v, align="center" if j in [1, 4, 5, 6, 7, 8, 9, 10] else "left")

        row += 1
        ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        _cell(ws2, row, 1, "TỔNG / TOTAL", bold=True, bg=YELLOW, align="center")
        totals = ["", "", "", "", total_vcpu, total_ram, total_os, total_data]
        for j, v in enumerate(totals, start=1):
            if j > 3:
                _cell(ws2, row, j, round(v, 1) if v else 0, bold=True, bg=YELLOW, align="center")

    # ── Sheet 3: Compute Sizing ────────────────────────────────────────────
    ws3 = wb.create_sheet("SIZING – COMPUTE")
    ws3.column_dimensions["A"].width = 45
    ws3.column_dimensions["B"].width = 20
    ws3.column_dimensions["C"].width = 15
    ws3.column_dimensions["D"].width = 40

    ws3.merge_cells("A1:D1")
    ws3["A1"].value = "SIZING KẾT QUẢ – COMPUTE / SERVER"
    ws3["A1"].font = Font(bold=True, size=14, color=WHITE)
    ws3["A1"].fill = PatternFill("solid", fgColor=DARK_BLUE)
    ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 35

    if ws_data and items:
        comp = sizing.calc_compute_sizing(ws_data, items)
        stor = sizing.calc_storage_sizing(ws_data, items)

        _section(ws3, 2, 1, "PHẦN 1 – THÔNG SỐ ĐẦU VÀO", 4)
        input_rows = [
            ("Tổng vCPU yêu cầu", comp["total_vcpu"], "vCPUs", ""),
            ("Tổng RAM yêu cầu", comp["total_ram_gb"], "GB", ""),
            ("Tỉ lệ ảo hóa CPU (vCPU:pCPU)", ws_data.virt_ratio, "", ""),
            ("CPU Overhead (%)", ws_data.cpu_overhead_pct, "%", ""),
            ("RAM Overhead (%)", ws_data.ram_overhead_pct, "%", ""),
            ("HA Reserve (%)", ws_data.ha_reserve_pct, "%", "N+1 = 25%"),
        ]
        for i, (label, val, unit, note) in enumerate(input_rows, start=3):
            _cell(ws3, i, 1, label, bg=LIGHT_GRAY)
            _cell(ws3, i, 2, val, align="center")
            _cell(ws3, i, 3, unit, align="center")
            _cell(ws3, i, 4, note)

        _section(ws3, 9, 1, "PHẦN 2 – TÍNH TOÁN SIZING SERVER", 4)
        calc_rows = [
            ("pCPU cores cần thiết (trước HA)", comp["pcpu_pre_ha"], "cores", "vCPU ÷ virt_ratio × (1+CPU_OH)"),
            ("pCPU cores sau HA reserve", comp["pcpu_with_ha"], "cores", "pCPU_pre × (1+HA%)"),
            ("RAM cần thiết (trước HA)", comp["ram_pre_ha_gb"], "GB", "RAM × (1+RAM_OH%)"),
            ("RAM sau HA reserve", comp["ram_with_ha_gb"], "GB", "RAM_pre × (1+HA%)"),
        ]
        for i, (label, val, unit, note) in enumerate(calc_rows, start=10):
            _cell(ws3, i, 1, label, bg=LIGHT_GRAY)
            _cell(ws3, i, 2, val, bold=True, align="center", bg="E2EFDA")
            _cell(ws3, i, 3, unit, align="center")
            _cell(ws3, i, 4, note)

        _section(ws3, 14, 1, "PHẦN 3 – CẤU HÌNH SERVER ĐỀ XUẤT", 4)
        cfg_rows = [
            ("CPU sockets / server", ws_data.cpu_sockets, "pcs", ""),
            ("Cores / socket", ws_data.cores_per_socket, "cores", ""),
            ("RAM / server", ws_data.ram_per_server_gb, "GB", ""),
            ("Số lượng server tối thiểu (compute)", comp["min_nodes"], "nodes", ""),
            ("Số lượng server sau HA (N+1)", comp["ha_nodes"], "nodes", ""),
            ("Dự phòng tăng trưởng (nodes thêm)", comp["growth_nodes"], "nodes", f"{ws_data.growth_years} năm @ {ws_data.growth_rate}%/năm"),
            ("⭐ TỔNG SERVER ĐỀ XUẤT", comp["total_nodes"], "nodes", ""),
        ]
        for i, (label, val, unit, note) in enumerate(cfg_rows, start=15):
            bold = i == 21
            bg_c = GREEN if i == 21 else None
            _cell(ws3, i, 1, label, bold=bold, bg=bg_c)
            _cell(ws3, i, 2, val, bold=bold, align="center", bg=bg_c)
            _cell(ws3, i, 3, unit, align="center")
            _cell(ws3, i, 4, note)

        _section(ws3, 22, 1, "PHẦN 4 – STORAGE SIZING", 4)
        st_rows = [
            ("Tổng Disk OS (GB)", stor["total_os_gb"], "GB", ""),
            ("Tổng Disk Data (GB)", stor["total_data_gb"], "GB", ""),
            ("Raw trước dedup (GB)", stor["raw_before_dedup_gb"], "GB", "OS+Data × (1+snap%) × (1+sys%)"),
            ("Usable sau dedup (GB)", stor["usable_after_dedup_gb"], "GB", f"÷ {ws_data.dedup_ratio}x dedup ratio"),
            ("Usable sau tăng trưởng (GB)", stor["usable_with_growth_gb"], "GB", f"{ws_data.growth_years} năm @ {ws_data.growth_rate}%/năm"),
            ("⭐ Usable cần thiết (TB)", stor["usable_tb"], "TB", ""),
            ("⭐ Raw cần thiết – RAID 5 (TB)", stor["raw_raid5_tb"], "TB", "75% efficiency"),
            ("⭐ Raw cần thiết – RAID 6 (TB)", stor["raw_raid6_tb"], "TB", "66% efficiency"),
            ("Tổng IOPS yêu cầu", stor["total_iops"], "IOPS", ""),
            ("Tổng Throughput yêu cầu", stor["total_throughput_mbps"], "MB/s", ""),
        ]
        for i, (label, val, unit, note) in enumerate(st_rows, start=23):
            bold = "⭐" in label
            bg_c = GREEN if bold else None
            _cell(ws3, i, 1, label, bold=bold, bg=bg_c)
            _cell(ws3, i, 2, val, bold=bold, align="center", bg=bg_c)
            _cell(ws3, i, 3, unit, align="center")
            _cell(ws3, i, 4, note)

    # ── Sheet 4: Backup ────────────────────────────────────────────────────
    ws4 = wb.create_sheet("BACKUP SURVEY & SIZING")
    ws4.column_dimensions["A"].width = 5
    ws4.column_dimensions["B"].width = 35
    ws4.column_dimensions["C"].width = 20
    for col in ["D", "E", "F", "G", "H"]:
        ws4.column_dimensions[col].width = 15

    ws4.merge_cells("A1:H1")
    ws4["A1"].value = "BACKUP SURVEY & SIZING – CHÍNH SÁCH & KẾT QUẢ SAO LƯU"
    ws4["A1"].font = Font(bold=True, size=14, color=WHITE)
    ws4["A1"].fill = PatternFill("solid", fgColor=DARK_BLUE)
    ws4["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws4.row_dimensions[1].height = 35

    if bk:
        _section(ws4, 2, 1, "PHẦN A – CHÍNH SÁCH BACKUP", 8)
        policy_rows = [
            ("Phần mềm Backup", bk.backup_software or ""),
            ("Backup Target", bk.backup_target or ""),
            ("Air-gap / Immutable", "Yes" if bk.air_gap_immutable else "No"),
            ("Offsite / Cloud Backup", "Yes" if bk.offsite_cloud else "No"),
            ("Cloud Target", bk.cloud_target or ""),
            ("Tape Backup", "Yes" if bk.tape_required else "No"),
        ]
        for i, (label, val) in enumerate(policy_rows, start=3):
            ws4.merge_cells(start_row=i, start_column=1, end_row=i, end_column=4)
            _cell(ws4, i, 1, label, bold=True, bg=LIGHT_GRAY)
            ws4.merge_cells(start_row=i, start_column=5, end_row=i, end_column=8)
            _cell(ws4, i, 5, val)

        _section(ws4, 9, 1, "PHẦN B – RTO / RPO THEO TIER", 8)
        _hdr(ws4, 10, 1, "Tier", size=9)
        ws4.merge_cells("B10:C10")
        _hdr(ws4, 10, 2, "Mô tả", size=9)
        _hdr(ws4, 10, 4, "RPO", size=9)
        _hdr(ws4, 10, 5, "RTO", size=9)
        ws4.merge_cells("F10:H10")
        _hdr(ws4, 10, 6, "Ví dụ ứng dụng", size=9)

        tiers = [
            ("Tier 1 – Critical", "Mission-critical", bk.tier1_rpo, bk.tier1_rto, "Core Banking, ERP"),
            ("Tier 2 – Important", "Ứng dụng quan trọng", bk.tier2_rpo, bk.tier2_rto, "Email, HR System"),
            ("Tier 3 – Standard", "Ứng dụng thông thường", bk.tier3_rpo, bk.tier3_rto, "Dev/Test, File Server"),
            ("Tier 4 – Archive", "Lưu trữ / Cold data", bk.tier4_rpo, bk.tier4_rto, "Long-term backup"),
        ]
        for i, (tier, desc, rpo, rto, ex) in enumerate(tiers, start=11):
            _cell(ws4, i, 1, tier, bold=True)
            ws4.merge_cells(start_row=i, start_column=2, end_row=i, end_column=3)
            _cell(ws4, i, 2, desc)
            _cell(ws4, i, 4, rpo, align="center")
            _cell(ws4, i, 5, rto, align="center")
            ws4.merge_cells(start_row=i, start_column=6, end_row=i, end_column=8)
            _cell(ws4, i, 6, ex)

        _section(ws4, 15, 1, "PHẦN C – NGUỒN DỮ LIỆU BACKUP", 8)
        src_headers = ["STT", "Nguồn dữ liệu", "Loại", "Dung lượng (TB)", "Tăng trưởng (%/năm)", "Tần suất", "Retention (ngày)", "Tier"]
        for j, h in enumerate(src_headers, start=1):
            _hdr(ws4, 16, j, h, size=9)
        for i, src in enumerate(sources, start=17):
            vals = [src.order_no, src.name, src.data_type, src.size_tb, src.growth_rate_pct, src.backup_frequency, src.retention_days, src.tier]
            for j, v in enumerate(vals, start=1):
                _cell(ws4, i, j, v, align="center" if j in [1, 4, 5, 7] else "left")

        row = 17 + len(sources)
        ws4.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        _cell(ws4, row, 1, "TỔNG DUNG LƯỢNG NGUỒN (TB)", bold=True, bg=YELLOW, align="center")
        _cell(ws4, row, 4, round(sum(s.size_tb for s in sources if s.size_tb), 2), bold=True, bg=YELLOW, align="center")

        if sources:
            bk_sizing = sizing.calc_backup_sizing(bk, sources)
            _section(ws4, row + 1, 1, "PHẦN D – KẾT QUẢ SIZING BACKUP REPOSITORY", 8)
            bk_rows = [
                ("Full backup (TB, trước dedup)", bk_sizing["full_backup_tb"], "TB"),
                ("Incremental backup (TB, trước dedup)", bk_sizing["incr_backup_tb"], "TB"),
                ("Tổng raw trước dedup (TB)", bk_sizing["raw_total_tb"], "TB"),
                ("Sau dedup/compress (TB)", bk_sizing["after_dedup_tb"], "TB"),
                ("⭐ Repository cần thiết (TB, incl. overhead)", bk_sizing["repo_needed_tb"], "TB"),
                ("Throughput tối thiểu (GB/h)", bk_sizing["min_throughput_gbph"], "GB/h"),
            ]
            for i, (label, val, unit) in enumerate(bk_rows, start=row + 2):
                bold = "⭐" in label
                bg_c = GREEN if bold else None
                ws4.merge_cells(start_row=i, start_column=1, end_row=i, end_column=5)
                _cell(ws4, i, 1, label, bold=bold, bg=bg_c)
                _cell(ws4, i, 6, val, bold=bold, align="center", bg=bg_c)
                _cell(ws4, i, 7, unit, align="center")

    # ── Sheet 5: Network ─────────────────────────────────────────────────────
    ws5 = wb.create_sheet("NETWORK & INFRA")
    ws5.column_dimensions["A"].width = 40
    ws5.column_dimensions["B"].width = 30
    ws5.column_dimensions["C"].width = 20

    ws5.merge_cells("A1:C1")
    ws5["A1"].value = "NETWORK & INFRASTRUCTURE SURVEY"
    ws5["A1"].font = Font(bold=True, size=14, color=WHITE)
    ws5["A1"].fill = PatternFill("solid", fgColor=DARK_BLUE)
    ws5["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws5.row_dimensions[1].height = 35

    if net:
        sections = [
            ("TOPOLOGY & SITE", [
                ("Số lượng site / datacenter", net.site_count),
                ("Mô hình triển khai", net.deployment_model),
                ("Khoảng cách giữa các site (km)", net.site_distance_km),
                ("Băng thông WAN (Gbps)", net.wan_bandwidth_gbps),
                ("Latency WAN (ms)", net.wan_latency_ms),
            ]),
            ("NETWORK FABRIC – COMPUTE", [
                ("Tốc độ kết nối server uplink", net.server_uplink_speed),
                ("Số lượng uplink / server", net.uplink_count_per_server),
                ("Top-of-Rack switch", net.tor_switch_status),
                ("Yêu cầu SDN / NSX", "Yes" if net.sdn_nsx_required else "No"),
                ("Yêu cầu RDMA / RoCE", "Yes" if net.rdma_roce_required else "No"),
            ]),
            ("STORAGE NETWORK", [
                ("Loại kết nối storage", net.storage_conn_type),
                ("Số lượng HBA/NIC / server", net.hba_nic_per_server),
                ("Fabric switch hiện có", "Yes" if net.fabric_switch_existing else "No"),
                ("FC ports cần thiết (tổng)", net.fc_ports_total),
                ("Yêu cầu multipath", "Yes" if net.multipath_required else "No"),
            ]),
            ("POWER & COOLING", [
                ("Công suất điện (kW) / rack", net.power_kw_per_rack),
                ("Số lượng rack dự kiến", net.rack_count),
                ("Redundant power (N+1)", "Yes" if net.redundant_power else "No"),
                ("Cooling type", net.cooling_type),
            ]),
            ("COMPLIANCE & SECURITY", [
                ("Yêu cầu tuân thủ", net.compliance_requirements),
                ("Mã hóa at-rest", "Yes" if net.encryption_at_rest else "No"),
                ("Mã hóa in-transit", "Yes" if net.encryption_in_transit else "No"),
                ("Air-gap / Isolated network", "Yes" if net.air_gap_required else "No"),
            ]),
        ]
        row = 2
        for section_title, fields in sections:
            _section(ws5, row, 1, section_title, 3)
            row += 1
            for label, val in fields:
                _cell(ws5, row, 1, label, bold=True, bg=LIGHT_GRAY)
                ws5.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
                _cell(ws5, row, 2, val if val is not None else "")
                row += 1

    # ── Sheet 6: Inventory ────────────────────────────────────────────────
    if inv:
        ws6 = wb.create_sheet("PHYSICAL INVENTORY")
        ws6.column_dimensions["A"].width = 5
        ws6.column_dimensions["B"].width = 30
        for col in ["C", "D", "E", "F", "G", "H"]:
            ws6.column_dimensions[col].width = 20

        ws6.merge_cells("A1:H1")
        ws6["A1"].value = "PHYSICAL INVENTORY – DANH MỤC THIẾT BỊ HIỆN HỮU"
        ws6["A1"].font = Font(bold=True, size=14, color=WHITE)
        ws6["A1"].fill = PatternFill("solid", fgColor=DARK_BLUE)
        ws6["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws6.row_dimensions[1].height = 35

        row = 2
        for cat_name, cat_data in [
            ("PHYSICAL SERVERS", inv.servers or []),
            ("SAN SWITCHES", inv.san_switches or []),
            ("STORAGE SYSTEMS", inv.storage_systems or []),
            ("NETWORK DEVICES", inv.network_devices or []),
            ("WIFI ACCESS POINTS", inv.wifi_aps or []),
        ]:
            _section(ws6, row, 1, cat_name, 8)
            row += 1
            hdrs = ["No.", "Tên thiết bị", "Model", "Vendor", "Serial Number", "SL", "Vị trí", "Trạng thái"]
            for j, h in enumerate(hdrs, start=1):
                _hdr(ws6, row, j, h, size=9)
            row += 1
            if cat_data:
                for i, item in enumerate(cat_data, start=1):
                    if isinstance(item, dict):
                        vals = [i, item.get("name", ""), item.get("model", ""), item.get("vendor", ""),
                                item.get("serial", ""), item.get("qty", 1), item.get("location", ""), item.get("status", "")]
                        for j, v in enumerate(vals, start=1):
                            _cell(ws6, row, j, v)
                        row += 1

    # ── Sheet 7: Application Inventory ───────────────────────────────────────
    apps = (app_inv.applications or []) if app_inv else []
    ws7a = wb.create_sheet("APPLICATION INVENTORY")
    ws7a.column_dimensions["A"].width = 5
    ws7a.column_dimensions["B"].width = 30
    ws7a.column_dimensions["C"].width = 15
    ws7a.column_dimensions["D"].width = 20
    ws7a.column_dimensions["E"].width = 20
    ws7a.column_dimensions["F"].width = 18
    ws7a.column_dimensions["G"].width = 28
    ws7a.column_dimensions["H"].width = 18
    ws7a.column_dimensions["I"].width = 20
    ws7a.column_dimensions["J"].width = 15
    ws7a.column_dimensions["K"].width = 18
    ws7a.column_dimensions["L"].width = 25

    ws7a.merge_cells("A1:L1")
    ws7a["A1"].value = "APPLICATION INVENTORY – DANH MỤC ỨNG DỤNG"
    ws7a["A1"].font = Font(bold=True, size=14, color=WHITE)
    ws7a["A1"].fill = PatternFill("solid", fgColor=DARK_BLUE)
    ws7a["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws7a.row_dimensions[1].height = 35

    app_hdrs = ["STT", "Tên ứng dụng", "Phiên bản", "Vendor", "Loại ứng dụng", "Môi trường",
                "Server/Host", "Database", "OS", "Criticality", "Hết hỗ trợ", "Ghi chú"]
    for j, h in enumerate(app_hdrs, start=1):
        _hdr(ws7a, 2, j, h, size=9)
    ws7a.row_dimensions[2].height = 30

    for i, app in enumerate(apps, start=3):
        if isinstance(app, dict):
            vals = [i - 2, app.get("name", ""), app.get("version", ""), app.get("vendor", ""),
                    app.get("app_type", ""), app.get("environment", ""), app.get("servers", ""),
                    app.get("database", ""), app.get("os", ""), app.get("criticality", ""),
                    app.get("support_expiry", ""), app.get("notes", "")]
            for j, v in enumerate(vals, start=1):
                _cell(ws7a, i, j, v, align="center" if j in [1, 10] else "left")

    if not apps:
        ws7a.merge_cells("A3:L3")
        _cell(ws7a, 3, 1, "Chưa có dữ liệu ứng dụng", align="center")

    # ── Sheet 8: Security ────────────────────────────────────────────────────
    if sec and sec.responses:
        ws7 = wb.create_sheet("SECURITY QUESTIONNAIRE")
        ws7.column_dimensions["A"].width = 5
        ws7.column_dimensions["B"].width = 15
        ws7.column_dimensions["C"].width = 55
        ws7.column_dimensions["D"].width = 50

        ws7.merge_cells("A1:D1")
        ws7["A1"].value = "SECURITY QUESTIONNAIRE – KHẢO SÁT BẢO MẬT"
        ws7["A1"].font = Font(bold=True, size=14, color=WHITE)
        ws7["A1"].fill = PatternFill("solid", fgColor=DARK_BLUE)
        ws7["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws7.row_dimensions[1].height = 35

        _hdr(ws7, 2, 1, "No.", size=9)
        _hdr(ws7, 2, 2, "Chủ đề", size=9)
        _hdr(ws7, 2, 3, "Câu hỏi", size=9)
        _hdr(ws7, 2, 4, "Phản hồi / Response", size=9)

        from schemas import SECURITY_QUESTIONS
        for i, q in enumerate(SECURITY_QUESTIONS, start=3):
            qid = str(q["id"])
            response = sec.responses.get(qid, "") if sec.responses else ""
            _cell(ws7, i, 1, q["id"], align="center")
            _cell(ws7, i, 2, q["topic"])
            _cell(ws7, i, 3, q["question"])
            _cell(ws7, i, 4, response, bg=YELLOW if not response else None)
            ws7.row_dimensions[i].height = 40

    # ── Sheet 8: OCP ─────────────────────────────────────────────────────────
    if ocp:
        ws8 = wb.create_sheet("SIZING – OCP")
        ws8.column_dimensions["A"].width = 35
        for col in ["B", "C", "D", "E", "F", "G", "H"]:
            ws8.column_dimensions[col].width = 16

        ws8.merge_cells("A1:H1")
        ws8["A1"].value = f"SIZING – RED HAT OPENSHIFT CONTAINER PLATFORM {ocp.ocp_version}"
        ws8["A1"].font = Font(bold=True, size=14, color=WHITE)
        ws8["A1"].fill = PatternFill("solid", fgColor=DARK_BLUE)
        ws8["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws8.row_dimensions[1].height = 35

        _section(ws8, 2, 1, "PHẦN A – CLUSTER PROFILE", 8)
        profile_rows = [
            ("Tên cluster / môi trường", ocp.cluster_name),
            ("Phiên bản OCP", ocp.ocp_version),
            ("Nền tảng hạ tầng", ocp.infra_platform),
            ("Deployment topology", ocp.deployment_topology),
            ("Network plugin", ocp.network_plugin),
            ("OpenShift Virtualization", "Yes" if ocp.ocp_virt_enabled else "No"),
            ("OpenShift Data Foundation (ODF)", "Yes" if ocp.odf_enabled else "No"),
            ("Tỉ lệ tăng trưởng (%/năm)", ocp.growth_rate_pct),
            ("Năm dự phòng sizing", ocp.sizing_years),
        ]
        for i, (label, val) in enumerate(profile_rows, start=3):
            _cell(ws8, i, 1, label, bold=True, bg=LIGHT_GRAY)
            ws8.merge_cells(start_row=i, start_column=2, end_row=i, end_column=8)
            _cell(ws8, i, 2, val)

        row = 12
        _section(ws8, row, 1, "PHẦN B – NODE SIZING", 8)
        row += 1
        node_hdrs = ["Loại Node", "Số lượng", "vCPU/Node", "RAM (GiB)/Node", "OS Disk (GB)", "Total vCPU", "Total RAM (GiB)", "Ghi chú"]
        for j, h in enumerate(node_hdrs, start=1):
            _hdr(ws8, row, j, h, size=9)
        row += 1

        ocp_result = sizing.calc_ocp_sizing(ocp)
        node_rows = [
            ("Control Plane (Master)", ocp.master_count, ocp.master_vcpu, ocp.master_ram_gib, ocp.master_disk_gb,
             ocp_result["master"]["total_vcpu"], ocp_result["master"]["total_ram_gib"], "Min: 3 nodes"),
            ("Worker Node", ocp.worker_count, ocp.worker_vcpu, ocp.worker_ram_gib, "-",
             ocp_result["worker"]["total_vcpu"], ocp_result["worker"]["total_ram_gib"], f"Với tăng trưởng: {ocp_result['worker']['workers_with_growth']} nodes"),
            ("Infrastructure Node", ocp.infra_count, ocp.infra_vcpu, ocp.infra_ram_gib, ocp.infra_disk_gb,
             ocp_result["infra"]["total_vcpu"], ocp_result["infra"]["total_ram_gib"], "Router, Registry, Monitoring"),
        ]
        if ocp.odf_enabled:
            node_rows.append(("ODF Storage Node", ocp.odf_node_count, ocp.odf_vcpu, ocp.odf_ram_gib, ocp.odf_disk_gb,
                               ocp_result["odf"]["total_vcpu"], ocp_result["odf"]["total_ram_gib"], "Ceph OSD, MON, MGR"))

        for vals in node_rows:
            for j, v in enumerate(vals, start=1):
                _cell(ws8, row, j, v, align="center" if j in [2, 3, 4, 5, 6, 7] else "left")
            row += 1

        _cell(ws8, row, 1, "⭐ CLUSTER TOTAL", bold=True, bg=GREEN)
        _cell(ws8, row, 6, ocp_result["cluster_total"]["vcpu"], bold=True, bg=GREEN, align="center")
        _cell(ws8, row, 7, ocp_result["cluster_total"]["ram_gib"], bold=True, bg=GREEN, align="center")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@router.get("/excel")
def export_excel(customer_id: int, db: Session = Depends(get_db)):
    c = db.query(models.Customer).filter(models.Customer.id == customer_id).first()
    if not c:
        raise HTTPException(status_code=404, detail="Customer not found")
    buf = build_excel(customer_id, db)
    filename = f"SVT_Survey_{c.name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
